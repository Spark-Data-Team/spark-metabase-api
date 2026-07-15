#!/usr/bin/env python3
"""Lit les 2 Excel de corpus (couleurs + blog/saisonnalité) : sheets, en-têtes, nb lignes,
échantillon, et détecte les mots-clés avec accents (pour le sujet GSC)."""
import sys
PATHS=[
 "/Users/louismonier/Downloads/Keywords couleurs Dash nanga Business.xlsx",
 "/Users/louismonier/Downloads/KW_Saisonnalite_Manucurist Blog.xlsx",
]
ACC="àâäéèêëîïôöùûüçÀÂÄÉÈÊËÎÏÔÖÙÛÜÇœæ"
def has_accent(s): return isinstance(s,str) and any(ch in s for ch in ACC)

try:
    from openpyxl import load_workbook
except ImportError:
    print("NO_OPENPYXL"); sys.exit(2)

for p in PATHS:
    print("\n================ ", p.split("/")[-1], " ================")
    try:
        wb=load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        print("  ERREUR ouverture:", repr(e)[:200]); continue
    print("  sheets:", wb.sheetnames)
    for ws in wb.worksheets:
        rows=list(ws.iter_rows(values_only=True))
        rows=[r for r in rows if any(c is not None for c in r)]  # drop empty rows
        print(f"\n  -- sheet '{ws.title}' : {len(rows)} lignes non vides --")
        if not rows: continue
        hdr=rows[0]
        print("     HEADERS:", hdr)
        for r in rows[1:13]:
            print("      ", r)
        if len(rows)>13: print(f"      ... (+{len(rows)-13} lignes)")
        acc=[r for r in rows[1:] if any(has_accent(c) for c in r)]
        print(f"     >>> lignes avec accents: {len(acc)} / {len(rows)-1}")
        for r in acc[:10]: print("        ACC:", r)
print("\nDONE")
